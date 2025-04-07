import openpyxl
path = "F:/Hexaware Training/Python/Pratice/FileHandling/ExcelHandling/one.xlsx"

# #To open a workbook
# wb_obj = openpyxl.load_workbook(path) # wb_obj is a work book object

# # To get a work book active sheet
# sheet_obj = wb_obj.active

# #Creating a cell object by using a sheet object
# cell_obj = sheet_obj.cell(row=2, column=2)
# print(cell_obj.value)


# writting in Excel file
wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row = 1, column = 1)
c1.value = "Aravind"
c2 = sheet.cell(row=1, column = 2)
c2.value = "S"

# or we can give based on index
c3 = sheet["A2"]
c3.value = "Ashwin"
c4 = sheet["B2"]
c4.value = "Harish"
wb.save(path)
print("Data Written")



